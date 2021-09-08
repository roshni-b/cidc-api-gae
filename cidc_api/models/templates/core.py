__all__ = [
    "Entry",
    "ExcelStyles",
    "MetadataTemplate",
    "MODEL_INSERTION_ORDER",
    "row_type_from_string",
    "RowType",
    "WorksheetConfig",
]

import datetime
from collections import defaultdict, OrderedDict
from enum import Enum
from warnings import filterwarnings
from typing import (
    Any,
    BinaryIO,
    Dict,
    Callable,
    List,
    Optional,
    OrderedDict as OrderedDict_Type,
    Tuple,
    Type,
    Union,
)

import openpyxl
from sqlalchemy import Column, Enum as SqlEnum
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell, xl_range

from .model_core import MetadataModel
from .utils import _all_bases, _get_global_insertion_order, insert_record_batch

MODEL_INSERTION_ORDER = _get_global_insertion_order()

filterwarnings(
    action="ignore",
    category=UserWarning,
    message="Data Validation extension is not supported",
    module="openpyxl",
)


class Entry:
    """
    One field in a metadata worksheet. Provides configuration for reading a
    value from a metadata spreadsheet cell into a sqlalchemy model attribute
    (or set of related attributes).

    Args:
      column: column attribute on a SQLAlchemy model class.
      name: optional human-readable label for this field, if not inferrable from `column`.
      process_as: optional dictionary mapping column attributes to data processing function.

    TO IMPLEMENT:
      encrypt: whether the spreadsheet value should be encrypted before storage. 
    """

    def __init__(
        self,
        column: Column,
        name: Optional[str] = None,
        gcs_uri_format: Optional[str] = None,
        process_as: Dict[Column, Callable[[str], Any]] = {},
    ):
        self.column = column
        self.name = name if name else column.name.replace("_", " ").capitalize()
        self.gcs_uri_format = gcs_uri_format
        self.process_as = process_as

        if gcs_uri_format and not hasattr(column.class_, "object_url"):
            raise Exception(
                f"gcs_uri_format should only be defined for columns on a File, not {column.class_.__name__}"
            )

        self.doc = column.doc
        self.sqltype = column.type
        self.enums = self.sqltype.enums if isinstance(self.sqltype, SqlEnum) else None
        self.pytype = self.sqltype.python_type

    def get_column_mapping(self, value, context: dict = {}) -> Dict[Column, Any]:
        """
        Given a value and a surround context to evalutate it in, returns a dict
        pointing table columns to their correctly-typed values. A single item
        if there's no process_as defined for this entry.
        Special handling for missing required values, date/time types, GCS URI
        formatting. Handling process_as is handled last.
        """
        if value is None:
            if self.column.nullable:
                return {self.column: None}
            else:
                raise Exception(f"Missing required value {self.name}")

        try:
            # Handle date/time parsing funkiness
            if self.pytype == datetime.time:
                try:
                    if not isinstance(value, datetime.datetime):
                        value = openpyxl.utils.datetime.from_excel(value)
                    processed_value = value.time()
                except:
                    raise TypeError(f"{value} is not a valid time")
            elif self.pytype == datetime.date:
                try:
                    if not isinstance(value, datetime.datetime):
                        value = openpyxl.utils.datetime.from_excel(value)
                    processed_value = value.date()
                except:
                    raise TypeError(f"{value} is not a valid date")

            else:
                processed_value = self.pytype(value)

            column_mapping = {self.column: processed_value}

            # Handle GCS URI formatting
            if self.gcs_uri_format:
                format_dict = {k.name: v for k, v in column_mapping.items()}
                format_dict.update({k.name: v for k, v in context.items()})
                column_mapping[
                    getattr(self.column.class_, "object_url")
                ] = self.gcs_uri_format.format(**format_dict)

            # Finally, handle process_as
            for column, process in self.process_as.items():
                context.update(column_mapping)
                column_mapping[column] = process(value, context)

        except Exception as e:
            raise Exception(
                f"Error in processing {self.name}={value}({type(value)}) as {self.pytype}: {e}"
            ) from e
        return column_mapping


class WorksheetConfig:
    """
    A worksheet within a metadata spreadsheet. A worksheet has a name, a list
    of `preamble_rows` and a dictionary mapping data column group headers to lists
    of `data_sections`.
    """

    name: str
    preamble: List[Entry]
    data_sections: Dict[str, List[Entry]]

    def __init__(
        self, name: str, preamble: List[Entry], data_sections: Dict[str, List[Entry]]
    ):
        self.name = name
        self.preamble = preamble
        self.data_sections = data_sections


class RowType(Enum):
    """Annotations denoting what type of data a template row contains."""

    TITLE = "#title"
    SKIP = "#skip"
    HEADER = "#header"
    PREAMBLE = "#preamble"
    DATA = "#data"


def row_type_from_string(maybe_type: str) -> Optional[RowType]:
    """Returns None if the input is not a valid RowType."""
    try:
        return RowType(maybe_type)
    except ValueError:
        return None


def _format_validation_range(
    validation_rows, validation_column, data_dict_worksheet_name
):
    start = xl_rowcol_to_cell(
        1,  # 1 is to skip first row in DD sheet that is for header
        validation_column,
        row_abs=True,
        col_abs=True,
    )
    stop = xl_rowcol_to_cell(
        validation_rows, validation_column, row_abs=True, col_abs=True
    )

    return f"'{data_dict_worksheet_name}'!{start}:{stop}"


class ExcelStyles:
    COLUMN_WIDTH = 30

    TITLE_STYLE_PROPS = {
        "border": 1,
        "bg_color": "#ffffb3",
        "bold": True,
        "align": "center",
        "text_wrap": True,
        "valign": "vcenter",
        "indent": 1,
    }
    PREAMBLE_STYLE_PROPS = {
        "border": 0,
        "top": 2,
        "top_color": "white",
        "bottom": 2,
        "bottom_color": "white",
        "bg_color": "#b2d2f6",
        "bold": True,
        "align": "right",
        "text_wrap": True,
        "valign": "vcenter",
        "indent": 1,
    }
    HEADER_STYLE_PROPS = {
        "border": 1,
        "bg_color": "#C6EFCE",
        "bold": True,
        "align": "center",
        "valign": "vcenter",
        "indent": 1,
    }
    DATA_STYLE_PROPS = {
        "border": 1,
        "bg_color": "#5fa3f0",
        "bold": True,
        "align": "center",
        "text_wrap": True,
        "valign": "vcenter",
        "indent": 1,
    }
    DIRECTIVE_STYLE_PROPS = {
        "border": 1,
        "bg_color": "#ffffb3",
        "bold": True,
        "align": "center",
        "text_wrap": True,
        "valign": "vcenter",
        "indent": 1,
    }
    COMMENT_STYLE_PROPS = {
        "color": "white",
        "font_size": 10,
        "x_scale": 2,
        "author": "CIDC",
    }

    def __init__(self, workbook: xlsxwriter.Workbook):
        self.TITLE_STYLE = workbook.add_format(self.TITLE_STYLE_PROPS)
        self.PREAMBLE_STYLE = workbook.add_format(self.PREAMBLE_STYLE_PROPS)
        self.HEADER_STYLE = workbook.add_format(self.HEADER_STYLE_PROPS)
        self.DATA_STYLE = workbook.add_format(self.DATA_STYLE_PROPS)
        self.DIRECTIVE_STYLE = workbook.add_format(self.DIRECTIVE_STYLE_PROPS)


class MetadataTemplate:
    """
    A metadata template. Must have attributes `upload_type`, `purpose`, and `worksheet_configs` defined.
    An optional `constants` dict allows for hidden values to be used later in reading the data, but does not
    affect the output template XLSX from `.write()`
    """

    upload_type: str
    purpose: str
    worksheet_configs: List[WorksheetConfig]
    constants: Dict[Column, Any]

    DATA_ROWS = 2000
    DATA_DICT_SHEETNAME = "Data Dictionary"

    def __init__(
        self,
        upload_type: str,
        purpose: str,
        worksheet_configs: List[WorksheetConfig],
        constants: Dict[Column, Any] = {},
    ):
        self.upload_type = upload_type
        self.purpose = purpose
        self.worksheet_configs = worksheet_configs
        self.constants = constants

    def read_and_insert(self, filename: Union[str, BinaryIO]) -> List[Exception]:
        """
        Extract the models from a populated instance of this template and try to
        insert them, rolling back and returning a list of errors if any are encountered.
        """
        try:
            records = self.read(filename)
        except Exception as e:
            return [e]
        else:
            return insert_record_batch(records)

    def read(
        self, filename: Union[str, BinaryIO]
    ) -> OrderedDict_Type[Type, List[MetadataModel]]:
        """
        Extract a list of SQLAlchemy models in insertion order from a populated
        instance of this template.
        """
        workbook = openpyxl.load_workbook(filename)

        # The preamble values flow across all sheets for context
        # (such as trial_id) to only need to collect each value once.
        # Constants also are used for all sheets.
        preamble_dict = self.constants.copy()

        # Extract partial model instances from the template
        model_instances: List[MetadataModel] = []
        model_dicts: List[Dict[Column, Any]] = []
        for config in self.worksheet_configs:
            if config.name not in workbook:
                raise Exception(
                    f"Missing expected worksheet {config.name}, please include even if all fields are optional"
                )

            # split the preamble and data rows
            preamble_rows = []
            data_rows = []
            data_header = None
            for n, row in enumerate(workbook[config.name].iter_rows()):
                # if no entries, skip it
                if all(cell.value is None for cell in row[1:]):
                    continue
                # row[0] is the type
                row_type: RowType = row_type_from_string(
                    row[0].value
                )  # None if not a valid type

                # only pay attention if we have a reason to care
                if row_type == RowType.PREAMBLE:
                    preamble_rows.append(row[1:])

                elif row_type == RowType.HEADER:
                    # if it's already been defined, that's an error
                    if data_header is not None:
                        raise Exception(
                            f"Second header row encountered at #{n+1} in worksheet {config.name}"
                        )
                    else:
                        data_header = row[1:]

                elif row_type == RowType.DATA:
                    if data_header is None:
                        raise Exception(
                            f"Encountered data row (#{n+1} in worksheet {config.name!r}) before header row"
                        )
                    else:
                        data_rows.append(row[1:])

                # ignore everything else
                else:
                    pass

            # turn the preamble into a mapping
            preamble_values = {
                str(row[0].value).lower(): row[1].value
                for row in preamble_rows
                if row[0].value is not None
            }
            # check the shape of the preamble
            if len(preamble_values) != len(config.preamble):
                raise Exception(
                    f"Expected {len(config.preamble)} preamble rows but saw {len(preamble_values)} in worksheet {config.name}"
                )
            for entry in config.preamble:
                # process the value
                try:
                    preamble_dict.update(
                        entry.get_column_mapping(
                            preamble_values.get(entry.name.lower())
                        )
                    )
                except Exception as e:
                    # add a bit of context
                    raise Exception(
                        f"Error in processing preamble {entry.name} in worksheet {config.name}: {e.__cause__ if e.__cause__ is not None else e}"
                    )

            model_dicts.append(preamble_dict)

            # combine and flatten the lists of configs across all data_sections
            data_configs = [
                entry for entries in config.data_sections.values() for entry in entries
            ]
            if len(data_rows):
                header_width = len([c for c in data_header if c.value is not None])
                # check the shape of the data
                if header_width != len(data_configs):
                    raise Exception(
                        f"Expected {len(data_configs)} data columns but saw {header_width} in worksheet {config.name}"
                    )

            for n, row in enumerate(data_rows):
                data_values = {
                    str(title_cell.value).lower(): data_cell.value
                    for title_cell, data_cell in zip(data_header, row)
                    if title_cell.value is not None
                }

                # context will be updated by every cell for each row,
                # but preamble_dict needs to persist unchanged
                context = preamble_dict.copy()
                data_dict = {}
                for entry in data_configs:
                    # process the value
                    try:
                        data_dict.update(
                            entry.get_column_mapping(
                                data_values.get(entry.name.lower()), context
                            )
                        )
                    except Exception as e:
                        # add some context here
                        raise Exception(
                            f"Error in processing {entry.name} for data row #{n+1} in worksheet {config.name}: {e.__cause__ if e.__cause__ is not None else e}"
                        )
                    context.update(data_dict)

                model_dicts.append(data_dict)

            # now let's take all our data from this sheet and make some models out of them
            for model_dict in model_dicts:
                # separate by target class, regardless of class hierarchy
                model_groups = defaultdict(dict)
                for column, value in model_dict.items():
                    model_groups[column.class_][column.name] = value

                # kwargs is {column.name: value}
                for model, kwargs in model_groups.items():
                    # also grab all the kwargs for all superclasses
                    [
                        kwargs.update(other_kwargs)
                        for other_model, other_kwargs in model_groups.items()
                        if other_model in _all_bases(model)
                    ]
                    model_instances.append(model(**kwargs))

        # Group model instances with matching values in their primary key or
        # unique-constrained columns.
        model_groups: Dict[
            Type[MetadataModel], Dict[Tuple, List[MetadataModel]]
        ] = defaultdict(lambda: defaultdict(list))
        for instance in model_instances:
            unique_values: Tuple = instance.unique_field_values()
            model_groups[instance.__class__][unique_values].append(instance)

        # Build a dictionary mapping model classes to deduplicated instances
        # of that class. Make sure that all instances have valid pk's for insertion,
        # remembering that some fk's can't be set until insert time
        deduped_instances: Dict[Type[MetadataModel], List[MetadataModel]] = defaultdict(
            list
        )
        for model, groups in model_groups.items():
            # keep track of which columns can be set later by fk
            # can only set if there's only a single foreign instance
            # also look at all the superclasses for hidden fk's
            fk_to_check = [fk for fk in model.__table__.foreign_keys]
            fk_to_check.extend(
                [
                    fk
                    for b in _all_bases(model)
                    if hasattr(b, "__table__")
                    for fk in b.__table__.foreign_keys
                ]
            )
            columns_to_set_later_by_fk = []
            for k, v in deduped_instances.items():
                for fk in fk_to_check:
                    if len(v) == 1 and fk.column.table.name in [k.__tablename__] + [
                        b.__tablename__
                        for b in _all_bases(k)
                        if hasattr(b, "__tablename__")
                    ]:
                        columns_to_set_later_by_fk.append(fk.parent)

            # special value None for all(unique_values is None)
            broad_instances = groups.pop(None, [])
            for specific_instances in groups.values():
                instance = model()
                for partial_instance in broad_instances + specific_instances:
                    instance.merge(partial_instance)

                for pre_col, pre_val in preamble_dict.items():
                    if (
                        hasattr(instance, pre_col.name)
                        and getattr(instance, pre_col.name) is None
                    ):
                        setattr(instance, pre_col.name, pre_val)

                if all(
                    [
                        # need to validate ALL of the primary keys
                        any(
                            [
                                # ANY of these is a reason that the value is fine
                                pk is not None,  # if it's set
                                any(
                                    [
                                        c.name == fk_col.name
                                        for fk_col in columns_to_set_later_by_fk
                                    ]
                                ),  # if it'll be set later
                                c.server_default is not None,  # if it has a default
                                c.autoincrement is True,  # if it autoincrements
                            ]
                        )
                        for c, pk in instance.primary_key_map().items()
                    ]
                ):
                    deduped_instances[model].append(instance)

        # now order the models for insertion based on the calculated order
        ordered_instances: OrderedDict_Type[Type, List[MetadataModel]] = OrderedDict()
        for next_model in MODEL_INSERTION_ORDER:
            next_instances = deduped_instances.pop(next_model, None)
            if next_instances:
                ordered_instances[next_model] = next_instances

        return ordered_instances

    def write(self, filename: str):
        """
        Generate an empty excel file for this template type.
        """
        workbook = xlsxwriter.Workbook(filename)
        styles = ExcelStyles(workbook)
        self._write_legend(workbook, styles)
        enum_ranges = self._write_data_dict(workbook, styles)
        self._write_worksheets(workbook, styles, enum_ranges)
        workbook.close()

    def _write_legend(self, workbook: xlsxwriter.Workbook, styles: ExcelStyles):
        ws = workbook.add_worksheet("Legend")
        ws.protect()
        ws.set_column(1, 100, width=styles.COLUMN_WIDTH)

        row = 0
        ws.write(row, 1, f"LEGEND", styles.DATA_STYLE)

        for config in self.worksheet_configs:
            row += 1
            ws.write(row, 1, f"Legend for tab {config.name!r}", styles.TITLE_STYLE)

            for entry in config.preamble:
                row += 1
                self._write_legend_item(ws, row, entry, styles.PREAMBLE_STYLE)

            for section_name, section_entries in config.data_sections.items():
                row += 1
                ws.write(
                    row,
                    1,
                    f"Section {section_name!r} of tab {config.name!r}",
                    styles.DIRECTIVE_STYLE,
                )

                for entry in section_entries:
                    row += 1
                    self._write_legend_item(ws, row, entry, styles.HEADER_STYLE)

    def _write_legend_item(
        self,
        ws: xlsxwriter.workbook.Worksheet,
        row: int,
        entry: Entry,
        style: xlsxwriter.workbook.Format,
    ):
        """ Writes a property with its type, description, and example if any."""
        ws.write(row, 1, entry.name, style)
        ws.write(row, 2, entry.sqltype.__class__.__name__)
        ws.write(row, 3, entry.doc)

    def _write_data_dict(
        self, workbook: xlsxwriter.Workbook, styles: ExcelStyles
    ) -> Dict[str, str]:
        ws = workbook.add_worksheet(self.DATA_DICT_SHEETNAME)
        ws.protect()
        ws.set_column(1, 100, width=styles.COLUMN_WIDTH)

        col = 0

        # a result dictionary that maps field names to data dictionary sheet
        # ranges of enum values to be used for validation
        data_dict_mapping = {}

        for config in self.worksheet_configs:
            for entry in config.preamble:
                rows = self._write_data_dict_item(ws, col, entry, styles.PREAMBLE_STYLE)
                if rows > 0:
                    # saving Data Dict range to use for validation
                    data_dict_mapping[entry.name] = _format_validation_range(
                        rows, col, self.DATA_DICT_SHEETNAME
                    )
                    col += 1

            for section_entries in config.data_sections.values():
                for entry in section_entries:
                    rows = self._write_data_dict_item(
                        ws, col, entry, styles.HEADER_STYLE
                    )
                    if rows > 0:
                        # saving Data Dict range to use for validation
                        data_dict_mapping[entry.name] = _format_validation_range(
                            rows, col, self.DATA_DICT_SHEETNAME
                        )
                        col += 1

        return data_dict_mapping

    def _write_data_dict_item(
        self,
        ws: xlsxwriter.workbook.Worksheet,
        col: int,
        entry: Entry,
        style: xlsxwriter.workbook.Format,
    ):
        if entry.enums is None:
            return 0

        # Write the data dict column header
        ws.write(0, col, entry.name, style)

        # Write the data dict column values
        for i, enum_value in enumerate(entry.enums):
            ws.write(1 + i, col, enum_value)

        return len(entry.enums)

    def _write_worksheets(
        self,
        workbook: xlsxwriter.Workbook,
        styles: ExcelStyles,
        enum_ranges: Dict[str, str],
    ):
        """Write content to the given worksheet"""
        for config in self.worksheet_configs:
            ws = workbook.add_worksheet(config.name)
            ws.set_column(0, 100, width=styles.COLUMN_WIDTH)

            row = 0
            col = 1

            # WORKSHEET TITLE
            self._write_type_annotation(ws, row, RowType.TITLE)
            preamble_range = xl_range(row, 1, row, 2)
            ws.merge_range(preamble_range, config.name, styles.TITLE_STYLE)

            # PREAMBLE ROWS
            row += 1
            for entry in config.preamble:
                # Write row type and entity name
                self._write_type_annotation(ws, row, RowType.PREAMBLE)
                ws.write(row, 1, entry.name, styles.PREAMBLE_STYLE)
                self._write_comment(ws, row, 1, entry, styles)

                # Format value cells next to entity name
                ws.write(row, 2, "", styles.PREAMBLE_STYLE)

                # Add data validation if appropriate
                value_cell = xl_rowcol_to_cell(row, 2)
                self._write_validation(ws, value_cell, entry, enum_ranges)
                row += 1

            # DATA SECTION HEADERS
            row += 1
            self._write_type_annotation(ws, row, RowType.SKIP)
            start_col = 1
            for section_header, section_entries in config.data_sections.items():
                section_width = len(section_entries)
                end_col = start_col + section_width - 1
                if end_col - start_col > 0:
                    ws.merge_range(
                        row,
                        start_col,
                        row,
                        end_col,
                        section_header,
                        styles.DIRECTIVE_STYLE,
                    )
                else:
                    ws.write(row, start_col, section_header, styles.DIRECTIVE_STYLE)
                start_col = end_col + 1

            # DATA SECTION TYPE ANNOTATIONS
            row += 1
            self._write_type_annotation(ws, row, RowType.HEADER)
            annotations = [RowType.DATA.value] * self.DATA_ROWS
            ws.write_column(row + 1, 0, annotations)

            # DATA SECTION ROWS
            for section_entries in config.data_sections.values():
                for entry in section_entries:
                    ws.write(row, col, entry.name, styles.HEADER_STYLE)
                    self._write_comment(ws, row, col, entry, styles)

                    # Write validation to data cells below header cell
                    data_range = xl_range(row + 1, col, row + self.DATA_ROWS, col)
                    self._write_validation(ws, data_range, entry, enum_ranges)
                    col += 1

    def _write_type_annotation(
        self, ws: xlsxwriter.workbook.Worksheet, row: int, row_type: RowType
    ):
        ws.write(row, 0, row_type.value)

    def _write_comment(
        self,
        ws: xlsxwriter.workbook.Worksheet,
        row: int,
        col: int,
        entry: Entry,
        styles: ExcelStyles,
    ):
        comment = entry.doc

        if entry.gcs_uri_format is not None:
            if isinstance(entry.gcs_uri_format, str):
                comment += f'\nIn .{entry.gcs_uri_format.split(".")[-1]} format'
            elif isinstance(entry.gcs_uri_format, dict):
                if "template_comment" in entry.gcs_uri_format:
                    comment += "\n" + entry.gcs_uri_format["template_comment"]

        if comment:
            ws.write_comment(row, col, comment, styles.COMMENT_STYLE_PROPS)

    def _write_validation(
        self,
        ws: xlsxwriter.workbook.Worksheet,
        cell_range: str,
        entry: Entry,
        enum_ranges: Dict[str, str],
    ):
        validation = self._get_validation(cell_range, entry, enum_ranges)
        if validation:
            ws.data_validation(cell_range, validation)

    def _get_validation(
        self, cell_range: str, entry: Entry, data_dict_validations: dict
    ) -> Optional[dict]:
        if entry.enums and len(entry.enums) > 0:
            data_dict_validation_range = data_dict_validations[entry.name]
            return {"validate": "list", "source": data_dict_validation_range}

        elif entry.pytype == datetime.date:
            return {
                "validate": "custom",
                "value": self._make_date_validation_string(cell_range),
                "error_message": "Please enter date in format mm/dd/yyyy",
            }
        elif entry.pytype == datetime.time:
            return {
                "validate": "time",
                "criteria": "between",
                "minimum": datetime.time(0, 0),
                "maximum": datetime.time(23, 59),
                "error_message": "Please enter time in format hh:mm",
            }
        elif entry.pytype == bool:
            return {"validate": "list", "source": ["True", "False"]}

        return None

    @staticmethod
    def _make_date_validation_string(cell_range: str) -> str:
        return f'=AND(ISNUMBER({cell_range}),LEFT(CELL("format",{cell_range}),1)="D")'
